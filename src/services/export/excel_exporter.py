"""Excel export service for quotes."""

import io
from typing import Dict, Any
from datetime import datetime
import boto3
from botocore.exceptions import ClientError

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from ...models.quote import Quote
from ...utils.logger import get_logger

logger = get_logger(__name__)


class ExcelExporter:
    """
    Service for exporting quotes to Excel format.
    
    Uses openpyxl to generate Excel workbooks with multiple sheets.
    Uploads to S3 and generates presigned URLs for download.
    """
    
    def __init__(self, s3_client=None, bucket_name: str = 'aws-pricing-quotes'):
        """
        Initialize the Excel exporter.
        
        Args:
            s3_client: Boto3 S3 client (optional, will create if not provided)
            bucket_name: S3 bucket name for storing Excel files
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError(
                "openpyxl is not installed. Install it with: pip install openpyxl"
            )
        
        self.s3_client = s3_client or boto3.client('s3')
        self.bucket_name = bucket_name
        logger.info(f"ExcelExporter initialized with bucket: {bucket_name}")
    
    def export_quote(
        self,
        quote: Quote,
        quote_content: Dict[str, Any],
        upload_to_s3: bool = True
    ) -> str:
        """
        Export quote to Excel format.
        
        Args:
            quote: Quote object
            quote_content: Structured quote content from QuoteGenerator
            upload_to_s3: Whether to upload to S3 (default: True)
        
        Returns:
            S3 presigned URL if uploaded, or local file path
        """
        logger.info(f"Exporting quote {quote.quote_id} to Excel")
        
        # Generate Excel in memory
        excel_buffer = io.BytesIO()
        self._generate_excel(excel_buffer, quote, quote_content)
        excel_buffer.seek(0)
        
        if upload_to_s3:
            # Upload to S3 and get presigned URL
            s3_key = f"quotes/{quote.user_id}/{quote.quote_id}.xlsx"
            url = self._upload_to_s3(excel_buffer, s3_key)
            logger.info(f"Excel uploaded to S3: {s3_key}")
            return url
        else:
            # Save to local file
            filename = f"quote_{quote.quote_id}.xlsx"
            with open(filename, 'wb') as f:
                f.write(excel_buffer.getvalue())
            logger.info(f"Excel saved locally: {filename}")
            return filename
    
    def _generate_excel(
        self,
        buffer: io.BytesIO,
        quote: Quote,
        content: Dict[str, Any]
    ) -> None:
        """Generate Excel workbook."""
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create sheets
        self._create_summary_sheet(wb, quote, content)
        self._create_original_services_sheet(wb, quote, content)
        self._create_aws_mappings_sheet(wb, quote, content)
        self._create_pricing_sheet(wb, quote, content)
        self._create_descriptions_sheet(wb, quote, content)
        
        # Save to buffer
        wb.save(buffer)
    
    def _create_summary_sheet(
        self,
        wb: Workbook,
        quote: Quote,
        content: Dict[str, Any]
    ) -> None:
        """Create summary sheet."""
        ws = wb.create_sheet("Summary", 0)
        
        # Styles
        header_font = Font(bold=True, size=14, color="FFFFFF")
        header_fill = PatternFill(start_color="232F3E", end_color="232F3E", fill_type="solid")
        title_font = Font(bold=True, size=18)
        
        # Title
        ws['A1'] = content['header']['title']
        ws['A1'].font = title_font
        ws.merge_cells('A1:B1')
        
        # Header information
        row = 3
        headers = [
            (self._get_translation(quote.language, 'quote_id'), quote.quote_id),
            (self._get_translation(quote.language, 'created_date'), content['header']['created_date']),
            (self._get_translation(quote.language, 'status'), content['header']['status']),
            (self._get_translation(quote.language, 'region'), content['header']['region'])
        ]
        
        for label, value in headers:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = value
            row += 1
        
        # Totals
        row += 1
        ws[f'A{row}'] = content['pricing']['total_monthly']['label']
        ws[f'A{row}'].font = Font(bold=True, size=12)
        ws[f'A{row}'].fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
        ws[f'B{row}'] = f"${content['pricing']['total_monthly']['value']:.2f} {content['pricing']['total_monthly']['currency']}"
        ws[f'B{row}'].font = Font(bold=True, size=12)
        ws[f'B{row}'].fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
        
        row += 1
        ws[f'A{row}'] = content['pricing']['total_annual']['label']
        ws[f'A{row}'].font = Font(bold=True, size=12)
        ws[f'A{row}'].fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
        ws[f'B{row}'] = f"${content['pricing']['total_annual']['value']:.2f} {content['pricing']['total_annual']['currency']}"
        ws[f'B{row}'].font = Font(bold=True, size=12)
        ws[f'B{row}'].fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
        
        # Notes
        if content['notes']:
            row += 2
            ws[f'A{row}'] = self._get_translation(quote.language, 'notes')
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
            ws[f'A{row}'] = content['notes']
            ws.merge_cells(f'A{row}:B{row}')
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 40
    
    def _create_original_services_sheet(
        self,
        wb: Workbook,
        quote: Quote,
        content: Dict[str, Any]
    ) -> None:
        """Create original services sheet."""
        ws = wb.create_sheet("Original Services")
        
        # Header
        headers = ['#', 'Provider', 'Service Name', 'Type', 'Specifications', 'Quantity', 'Region']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="232F3E", end_color="232F3E", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Data
        for i, service in enumerate(content['original_services']['services'], 1):
            row = i + 1
            ws.cell(row=row, column=1, value=i)
            ws.cell(row=row, column=2, value=service['provider'])
            ws.cell(row=row, column=3, value=service['service_name'])
            ws.cell(row=row, column=4, value=service['service_type'])
            ws.cell(row=row, column=5, value=self._format_dict(service['specifications']))
            ws.cell(row=row, column=6, value=service['quantity'])
            ws.cell(row=row, column=7, value=service.get('region', ''))
        
        # Adjust column widths
        for col in range(1, 8):
            ws.column_dimensions[get_column_letter(col)].width = 20
        ws.column_dimensions['E'].width = 40  # Specifications column
    
    def _create_aws_mappings_sheet(
        self,
        wb: Workbook,
        quote: Quote,
        content: Dict[str, Any]
    ) -> None:
        """Create AWS mappings sheet."""
        ws = wb.create_sheet("AWS Mappings")
        
        # Header
        headers = ['#', 'AWS Service', 'Category', 'Type', 'Specifications', 'Confidence', 'Explanation', 'Alternatives']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="232F3E", end_color="232F3E", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Data
        for i, mapping in enumerate(content['aws_mappings']['mappings'], 1):
            row = i + 1
            ws.cell(row=row, column=1, value=i)
            ws.cell(row=row, column=2, value=mapping['aws_service'])
            ws.cell(row=row, column=3, value=mapping['aws_service_category'])
            ws.cell(row=row, column=4, value=mapping['aws_service_type'])
            ws.cell(row=row, column=5, value=self._format_dict(mapping['specifications']))
            ws.cell(row=row, column=6, value=f"{mapping['confidence_score']:.2f}")
            ws.cell(row=row, column=7, value=mapping['explanation'])
            ws.cell(row=row, column=8, value=', '.join(mapping['alternatives']))
        
        # Adjust column widths
        for col in range(1, 9):
            ws.column_dimensions[get_column_letter(col)].width = 20
        ws.column_dimensions['E'].width = 40  # Specifications column
        ws.column_dimensions['G'].width = 50  # Explanation column
    
    def _create_pricing_sheet(
        self,
        wb: Workbook,
        quote: Quote,
        content: Dict[str, Any]
    ) -> None:
        """Create pricing sheet."""
        ws = wb.create_sheet("Pricing")
        
        # Header
        headers = ['#', 'Region', 'Pricing Model', 'Monthly Cost', 'Annual Cost', 'Currency', 'Available']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="232F3E", end_color="232F3E", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Data
        for i, item in enumerate(content['pricing']['items'], 1):
            row = i + 1
            ws.cell(row=row, column=1, value=i)
            ws.cell(row=row, column=2, value=item['region'])
            ws.cell(row=row, column=3, value=item['pricing_model'])
            ws.cell(row=row, column=4, value=item['monthly_cost'])
            ws.cell(row=row, column=4).number_format = '$#,##0.00'
            ws.cell(row=row, column=5, value=item['annual_cost'])
            ws.cell(row=row, column=5).number_format = '$#,##0.00'
            ws.cell(row=row, column=6, value=item['currency'])
            ws.cell(row=row, column=7, value='Yes' if item['region_availability'] else 'No')
        
        # Totals
        row = len(content['pricing']['items']) + 3
        ws.cell(row=row, column=3, value=content['pricing']['total_monthly']['label'])
        ws.cell(row=row, column=3).font = Font(bold=True)
        ws.cell(row=row, column=3).fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
        ws.cell(row=row, column=4, value=content['pricing']['total_monthly']['value'])
        ws.cell(row=row, column=4).number_format = '$#,##0.00'
        ws.cell(row=row, column=4).font = Font(bold=True)
        ws.cell(row=row, column=4).fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
        
        row += 1
        ws.cell(row=row, column=3, value=content['pricing']['total_annual']['label'])
        ws.cell(row=row, column=3).font = Font(bold=True)
        ws.cell(row=row, column=3).fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
        ws.cell(row=row, column=4, value=content['pricing']['total_annual']['value'])
        ws.cell(row=row, column=4).number_format = '$#,##0.00'
        ws.cell(row=row, column=4).font = Font(bold=True)
        ws.cell(row=row, column=4).fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
        
        # Adjust column widths
        for col in range(1, 8):
            ws.column_dimensions[get_column_letter(col)].width = 18
    
    def _create_descriptions_sheet(
        self,
        wb: Workbook,
        quote: Quote,
        content: Dict[str, Any]
    ) -> None:
        """Create descriptions and disclaimers sheet."""
        ws = wb.create_sheet("Info & Disclaimers")
        
        row = 1
        
        # Service Descriptions
        if content['descriptions']['services']:
            ws[f'A{row}'] = content['descriptions']['title']
            ws[f'A{row}'].font = Font(bold=True, size=14)
            row += 2
            
            for desc in content['descriptions']['services']:
                ws[f'A{row}'] = desc['service']
                ws[f'A{row}'].font = Font(bold=True)
                ws[f'B{row}'] = desc['description']
                row += 1
            
            row += 1
        
        # Benefits
        ws[f'A{row}'] = content['benefits']['title']
        ws[f'A{row}'].font = Font(bold=True, size=14)
        row += 2
        
        for benefit in content['benefits']['items']:
            ws[f'A{row}'] = f"• {benefit}"
            ws.merge_cells(f'A{row}:B{row}')
            row += 1
        
        row += 1
        
        # Disclaimers
        ws[f'A{row}'] = content['disclaimers']['title']
        ws[f'A{row}'].font = Font(bold=True, size=14)
        row += 2
        
        for disclaimer in content['disclaimers']['items']:
            ws[f'A{row}'] = f"• {disclaimer}"
            ws.merge_cells(f'A{row}:B{row}')
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 80
    
    def _upload_to_s3(self, buffer: io.BytesIO, s3_key: str) -> str:
        """
        Upload Excel to S3 and generate presigned URL.
        
        Args:
            buffer: Excel buffer
            s3_key: S3 object key
        
        Returns:
            Presigned URL for download
        """
        try:
            # Upload to S3
            self.s3_client.put_object(
                Bucket=self.bucket_name,
                Key=s3_key,
                Body=buffer.getvalue(),
                ContentType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                ServerSideEncryption='AES256'
            )
            
            # Generate presigned URL (valid for 7 days)
            url = self.s3_client.generate_presigned_url(
                'get_object',
                Params={
                    'Bucket': self.bucket_name,
                    'Key': s3_key
                },
                ExpiresIn=604800  # 7 days
            )
            
            return url
        except ClientError as e:
            logger.error(f"Failed to upload Excel to S3: {e}")
            raise
    
    def _format_dict(self, d: Dict[str, Any]) -> str:
        """Format dictionary as string."""
        items = [f"{k}: {v}" for k, v in d.items()]
        return ", ".join(items)
    
    def _get_translation(self, language: str, key: str) -> str:
        """Get translation for a key."""
        translations = {
            'en': {
                'quote_id': 'Quote ID',
                'created_date': 'Created Date',
                'status': 'Status',
                'region': 'Region',
                'notes': 'Notes'
            },
            'zh': {
                'quote_id': '报价单编号',
                'created_date': '创建日期',
                'status': '状态',
                'region': '区域',
                'notes': '备注'
            }
        }
        
        return translations.get(language, translations['en']).get(key, key)
